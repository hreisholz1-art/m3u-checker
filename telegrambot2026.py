import os
import tempfile
import zipfile
import logging
import asyncio
import sqlite3
from pathlib import Path
from datetime import datetime
from contextlib import asynccontextmanager
import json
import base64
import traceback
import re

from fastapi import FastAPI, Request, HTTPException
from telegram import Update
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
)
from dotenv import load_dotenv
import requests

# Google Sheets
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ─────────────── ЛОГИРОВАНИЕ ───────────────
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# ─────────────── КОНФИГ ───────────────
load_dotenv()
BOT_TOKEN = os.getenv("BOT_TOKEN")
WEBHOOK_SECRET = os.getenv("WEBHOOK_SECRET", "change-me-very-secure-secret-2026")
COMBINER_SCRIPT = "m3u_combiner_fixed.py"

if not BOT_TOKEN:
    raise ValueError("BOT_TOKEN не задан")

application: Application = None

# ─────────────── ЛОКАЛЬНАЯ БД ───────────────
DB_PATH = Path("dividends.db")

def init_db():
    """Инициализация SQLite базы данных"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS dividends (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            wkn TEXT NOT NULL,
            name TEXT NOT NULL,
            amount REAL NOT NULL,
            logo_url TEXT,
            year INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS wkn_lookup (
            code TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            logo_url TEXT
        )
    """)
    conn.commit()
    conn.close()
    logger.info("✅ База данных инициализирована")

def load_wkn_json():
    """Загрузка WKN данных из JSON в SQLite"""
    try:
        if not Path("wkn.json.txt").exists():
            logger.warning("wkn.json.txt не найден - пропускаем загрузку")
            return
            
        with open("wkn.json.txt", "r", encoding="utf-8") as f:
            data = json.load(f)
        
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        count = 0
        for item in data:
            name = item.get("name", "")
            logo = item.get("logo_url", "").strip()
            wkn = item.get("wkn", "").strip().upper()
            isin = item.get("isin", "").strip().upper()
            
            if wkn:
                c.execute("INSERT OR REPLACE INTO wkn_lookup (code, name, logo_url) VALUES (?, ?, ?)",
                         (wkn, name, logo))
                count += 1
            if isin:
                c.execute("INSERT OR REPLACE INTO wkn_lookup (code, name, logo_url) VALUES (?, ?, ?)",
                         (isin, name, logo))
                count += 1
        
        conn.commit()
        conn.close()
        logger.info(f"✅ Загружено {count} записей из wkn.json.txt")
    except Exception as e:
        logger.error(f"Ошибка загрузки wkn.json.txt: {e}")

def get_wkn_info(code: str):
    """Получить информацию о WKN/ISIN из базы"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT name, logo_url FROM wkn_lookup WHERE code = ?", (code.upper(),))
    result = c.fetchone()
    conn.close()
    
    if result:
        return {"name": result[0], "logo": result[1]}
    return None

def add_dividend_to_db(date: str, wkn: str, name: str, amount: float, logo_url: str = "", year: int = None):
    """Добавить запись о дивиденде в локальную БД"""
    if year is None:
        year = datetime.now().year
    
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT INTO dividends (date, wkn, name, amount, logo_url, year)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (date, wkn, name, amount, logo_url, year))
    conn.commit()
    conn.close()

def delete_dividends_by_date(date: str, year: int = None):
    """Удалить записи по дате из БД"""
    if year is None:
        year = datetime.now().year
    
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM dividends WHERE date = ? AND year = ?", (date, year))
    deleted = c.rowcount
    conn.commit()
    conn.close()
    return deleted

def generate_excel(year: int = None):
    """Генерация Excel файла с дивидендами"""
    if year is None:
        year = datetime.now().year
    
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT date, logo_url, wkn, name, amount 
        FROM dividends 
        WHERE year = ? 
        ORDER BY date
    """, (year,))
    rows = c.fetchall()
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = str(year)
    
    # Заголовки
    headers = ["Дата", "Логотип", "WKN", "Акция", "Сумма (€)"]
    ws.append(headers)
    
    # Цвета для разных WKN
    colors = [
        "FFCCCC", "CCFFCC", "CCCCFF", "FFFFCC", "CCFFFF",
        "FFCCFF", "FFE6CC", "E6CCFF", "CCE6FF", "FFCCAA"
    ]
    wkn_colors = {}
    
    # Данные
    current_row = 2
    for row_data in rows:
        ws.append(row_data)
        current_row += 1
        
        # Цвет для WKN
        wkn = row_data[2]
        if wkn not in wkn_colors:
            wkn_colors[wkn] = colors[len(wkn_colors) % len(colors)]
        
        fill = PatternFill(start_color=wkn_colors[wkn], end_color=wkn_colors[wkn], fill_type="solid")
        for col in range(1, 6):
            ws.cell(row=current_row, column=col).fill = fill
    
    # Строка с суммой
    sum_row = current_row + 1
    ws[f"A{sum_row}"] = "ИТОГО"
    ws[f"E{sum_row}"] = f"=SUM(E2:E{current_row})"
    
    # Форматирование
    for col in range(1, 6):
        ws.cell(row=sum_row, column=col).fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
    
    # Ширина колонок
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 12
    
    # Сохранение
    output_path = Path(f"dividends_{year}.xlsx")
    wb.save(output_path)
    return output_path

# ─────────────── GOOGLE SHEETS ───────────────
COLORS = [
    {"red": 1.0, "green": 0.9, "blue": 0.9},
    {"red": 0.9, "green": 1.0, "blue": 0.9},
    {"red": 0.9, "green": 0.9, "blue": 1.0},
    {"red": 1.0, "green": 1.0, "blue": 0.9},
    {"red": 0.9, "green": 1.0, "blue": 1.0},
]

def get_color_for_wkn(wkn: str):
    return COLORS[hash(wkn) % len(COLORS)]

def _get_spreadsheet():
    b64 = os.getenv("GOOGLE_CREDENTIALS_BASE64")
    if not b64:
        raise ValueError("GOOGLE_CREDENTIALS_BASE64 не задан")
    creds_dict = json.loads(base64.b64decode(b64).decode('utf-8'))
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
    client = gspread.authorize(creds)
    return client.open_by_key("1r2P4pF1TcICCuUAZNZm5lEpykVVZe94QZQ6-z6CrNg8")

def add_dividend_to_sheets(date: str, wkn: str, name: str, amount: float, logo_url: str = ""):
    """Добавить дивиденд в Google Sheets"""
    try:
        sheet = _get_spreadsheet().sheet1
        rows = sheet.get_all_values()
        last_row = len(rows)
        if last_row < 2:
            last_row = 2

        data_row = last_row + 1
        sum_row = data_row + 1

        sheet.update(f"A{data_row}", [[date, '', name, wkn, amount]])        
        # Цвет
        color = get_color_for_wkn(wkn)
        sheet.format(f"A{data_row}:E{data_row}", {"backgroundColor": color})
        
        # Формула суммы
        sheet.update(f"E{sum_row}", f"=SUM(E3:E{data_row})")
        
        return True
    except Exception as e:
        logger.error(f"Ошибка добавления в Google Sheets: {e}")
        return False

def delete_from_sheets(date: str):
    """Удалить записи из Google Sheets по дате"""
    try:
        sheet = _get_spreadsheet().sheet1
        rows = sheet.get_all_values()
        to_del = [i+1 for i, r in enumerate(rows[2:], start=3) if r and r[0] == date]
        
        for i in sorted(to_del, reverse=True):
            sheet.delete_rows(i)
        
        # Обновить формулу суммы
        last = max(3, len(sheet.get_all_values()))
        sheet.update("E2", f"=SUM(E3:E{last})")
        
        return len(to_del)
    except Exception as e:
        logger.error(f"Ошибка удаления из Google Sheets: {e}")
        return 0

# ─────────────── TELEGRAM ХЕНДЛЕРЫ ───────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 Привет!\n\n"
        "📺 Пришли .m3u/.m3u8/.txt файл — проверю потоки\n"
        "💰 Скрытые команды: /mysecret"
    )

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    document = update.message.document
    if not document:
        await update.message.reply_text("Пришли файл.")
        return

    name = (document.file_name or "").lower()
    if not any(name.endswith(ext) for ext in ('.m3u', '.m3u8', '.txt', '.text')):
        await update.message.reply_text("Поддерживаются: .m3u, .m3u8, .txt")
        return

    msg = await update.message.reply_text("📥 Скачиваю...")
    
    try:
        with tempfile.TemporaryDirectory() as tmp_str:
            tmp = Path(tmp_str)
            input_path = tmp / "input.m3u"
            
            # Скачивание файла
            file = await document.get_file()
            await file.download_to_drive(custom_path=str(input_path))

            await msg.edit_text("🔍 Проверяю потоки... (3–20 мин)")

            # Проверка FFmpeg
            try:
                proc = await asyncio.create_subprocess_exec(
                    "ffmpeg", "-version",
                    stdout=asyncio.subprocess.DEVNULL,
                    stderr=asyncio.subprocess.DEVNULL
                )
                await proc.communicate()
                if proc.returncode != 0:
                    raise FileNotFoundError("FFmpeg error")
            except Exception as e:
                logger.error(f"FFmpeg check failed: {e}")
                await msg.edit_text("❌ FFmpeg не установлен на сервере")
                return

            output_m3u = tmp / "good.m3u"
            
            # Запуск скрипта проверки
            cmd = [
                "python3", COMBINER_SCRIPT, 
                str(tmp), 
                "-w", "4", 
                "-t", "15", 
                "-o", str(output_m3u)
            ]
            
            proc = await asyncio.create_subprocess_exec(
                *cmd,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE
            )
            
            stdout, stderr = await proc.communicate()
            
            if proc.returncode != 0:
                error_msg = stderr.decode('utf-8', errors='ignore')[:500]
                logger.error(f"Combiner error: {error_msg}")
                await msg.edit_text(f"❌ Ошибка обработки:\n{error_msg}")
                return

            if not output_m3u.exists() or output_m3u.stat().st_size < 200:
                await msg.edit_text("❌ Не найдено рабочих потоков")
                return

            # Создание ZIP
            zip_name = f"m3u_{datetime.now().strftime('%Y%m%d_%H%M')}.zip"
            zip_path = tmp / zip_name
            
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                zf.write(output_m3u, "good.m3u")

            # Отправка
            file_size = zip_path.stat().st_size
            if file_size > 50 * 1024 * 1024:
                await msg.edit_text("❌ Файл >50 МБ — слишком большой для отправки")
                return
            
            await msg.edit_text("📤 Отправляю результат...")
            
            with open(zip_path, "rb") as f:
                await update.message.reply_document(
                    document=f,
                    filename=zip_name,
                    caption=f"✅ Готово! Размер: {file_size / 1024:.1f} KB"
                )
            
            await msg.delete()

    except Exception as e:
        logger.exception("Критическая ошибка при обработке плейлиста")
        try:
            await msg.edit_text(f"💥 Ошибка: {str(e)[:200]}")
        except:
            await update.message.reply_text("💥 Ошибка обработки файла")

# ─────────────── СКРЫТЫЕ КОМАНДЫ ───────────────
async def handle_hidden_commands(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()

    # /mysecret - показать команды
    if text == "/mysecret":
        await update.message.reply_text(
            "🔐 Скрытые команды:\n\n"
            "• <code>wkn123456 45.50euro</code> — добавить дивиденд\n"
            "• <code>isinDE00012345 30euro</code> — добавить по ISIN\n"
            "• <code>del02.06</code> — удалить записи за 2 июня\n"
            "• <code>new27</code> — создать лист на 2027 год\n"
            "• <code>/divxlsx</code> — скачать Excel\n"
            "• <code>/divlog</code> — последние записи\n"
            "• <code>/divdebug</code> — тест Google Sheets",
            parse_mode="HTML"
        )
        return

    # new27 - создать новый год в Sheets
    if match := re.fullmatch(r"new(\d{2})", text, re.IGNORECASE):
        year = f"20{match.group(1)}"
        try:
            sh = _get_spreadsheet()
            sh.duplicate_sheet(sh.sheet1.id, insert_sheet_index=1, new_sheet_name=year)
            sheet = sh.worksheet(year)
            sheet.clear()
            sheet.update("A1:E2", [
                ["Дата", "Логотип", "WKN", "Акция", "Сумма (€)"],
                ["", "", "", "", "=SUM(E3:E1000)"]
            ])
            await update.message.reply_text(f"🆕 Лист {year} создан в Google Sheets")
        except Exception as e:
            logger.error(f"new error: {e}")
            await update.message.reply_text("❌ Ошибка создания листа")
        return

    # del02.06 - удалить записи по дате
    if match := re.fullmatch(r"del(\d{2})\.(\d{2})", text, re.IGNORECASE):
        day, month = match.groups()
        target = f"{day}.{month}.{datetime.now().year}"
        
        # Удалить из БД
        deleted_db = delete_dividends_by_date(target)
        
        # Удалить из Sheets
        deleted_sheets = delete_from_sheets(target)
        
        await update.message.reply_text(
            f"🗑️ Удалено:\n"
            f"📊 БД: {deleted_db}\n"
            f"📈 Sheets: {deleted_sheets}"
        )
        return

    # wkn123456 45.50euro или isin... 30euro
    match = re.fullmatch(
        r"(?P<prefix>wkn|isin)(?P<code>[a-zA-Z0-9]{6,12})\s+(?P<amount>\d+\.?\d*)\s*euro",
        text,
        re.IGNORECASE
    )
    if match:
        code = match.group("code").upper()
        amount = float(match.group("amount"))

        try:
            # Поиск в базе
            stock_info = get_wkn_info(code)
            
            if stock_info:
                stock_name = stock_info["name"]
                logo_url = stock_info["logo"]
            else:
                stock_name = f"WKN{code}"
                logo_url = ""

            date_str = datetime.now().strftime("%d.%m.%Y")
            year = datetime.now().year
            
            # Добавить в БД
            add_dividend_to_db(date_str, code, stock_name, amount, logo_url, year)
            
            # Добавить в Google Sheets
            sheets_ok = add_dividend_to_sheets(date_str, code, stock_name, amount, logo_url)
            
            status = "✅ Добавлено в БД и Sheets" if sheets_ok else "⚠️ Добавлено в БД (Sheets недоступен)"
            
            await update.message.reply_text(
                f"{status}\n"
                f"📅 {date_str}\n"
                f"🏢 {stock_name}\n"
                f"💶 {amount}€"
            )
        except Exception as e:
            logger.error(f"Ошибка добавления дивиденда: {e}")
            await update.message.reply_text("❌ Ошибка при добавлении")
        return

# /divxlsx - скачать Excel
async def download_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        year = datetime.now().year
        xlsx_path = generate_excel(year)
        
        with open(xlsx_path, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename=f"dividends_{year}.xlsx",
                caption=f"📊 Дивиденды за {year} год"
            )
        
        xlsx_path.unlink()  # Удалить временный файл
    except Exception as e:
        logger.error(f"Ошибка генерации Excel: {e}")
        await update.message.reply_text("❌ Ошибка при создании файла")

# /divlog - показать последние записи
async def show_log(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("""
            SELECT date, wkn, name, amount 
            FROM dividends 
            WHERE year = ? 
            ORDER BY created_at DESC 
            LIMIT 10
        """, (datetime.now().year,))
        rows = c.fetchall()
        conn.close()
        
        if not rows:
            await update.message.reply_text("📭 Записей нет")
            return
        
        text = "📋 Последние 10 записей:\n\n"
        total = 0
        for date, wkn, name, amount in rows:
            text += f"• {date} | {wkn} | {name} | {amount}€\n"
            total += amount
        
        text += f"\n💰 Сумма: {total:.2f}€"
        await update.message.reply_text(text)
        
    except Exception as e:
        logger.error(f"Ошибка показа логов: {e}")
        await update.message.reply_text("❌ Ошибка")

# /divdebug - диагностика Google Sheets
async def divdebug(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        b64 = os.getenv("GOOGLE_CREDENTIALS_BASE64")
        if not b64:
            await update.message.reply_text("❌ GOOGLE_CREDENTIALS_BASE64 не задан")
            return

        creds_dict = json.loads(base64.b64decode(b64).decode('utf-8'))
        if "client_email" not in creds_dict:
            await update.message.reply_text("❌ Неверный формат credentials.json")
            return

        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        client = gspread.authorize(creds)
        sheet = client.open_by_key("1r2P4pF1TcICCuUAZNZm5lEpykVVZe94QZQ6-z6CrNg8").sheet1
        value = sheet.acell("A1").value or "пусто"

        await update.message.reply_text(
            f"✅ Google Sheets подключен!\n"
            f"Email: {creds_dict['client_email']}\n"
            f"A1: {value}"
        )

    except Exception as e:
        error_detail = traceback.format_exc()
        msg = f"❌ Ошибка:\n\n{error_detail[-3900:]}"
        await update.message.reply_text(msg)

# ─────────────── FASTAPI ───────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    global application
    
    # Инициализация
    init_db()
    load_wkn_json()
    
    application = Application.builder().token(BOT_TOKEN).build()
    await application.initialize()
    await application.start()
    
    # Handlers
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("mysecret", handle_hidden_commands))
    application.add_handler(CommandHandler("divxlsx", download_excel))
    application.add_handler(CommandHandler("divlog", show_log))
    application.add_handler(CommandHandler("divdebug", divdebug))
    application.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_hidden_commands))
    
    logger.info("✅ Бот запущен")
    yield
    
    # Shutdown
    await application.stop()
    await application.shutdown()

app = FastAPI(title="M3U + Dividends Bot 2026", lifespan=lifespan)

@app.get("/")
async def root():
    return {"status": "running", "bot": "M3U + Dividends", "version": "2026.1"}

@app.get("/health")
async def health():
    return {"status": "ok"}

@app.post(f"/webhook/{WEBHOOK_SECRET}")
async def webhook(request: Request):
    if request.headers.get("X-Telegram-Bot-Api-Secret-Token") != WEBHOOK_SECRET:
        raise HTTPException(403, "Forbidden")
    
    try:
        update = Update.de_json(await request.json(), application.bot)
        await application.update_queue.put(update)
        return {"ok": True}
    except Exception as e:
        logger.error("Webhook error", exc_info=True)
        raise HTTPException(500, str(e))

if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
